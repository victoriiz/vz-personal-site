#!/usr/bin/env python3
"""
Convert books.xlsx (the hand-maintained reading spreadsheet) into books.json.

The spreadsheet is the source of truth you edit; the site reads books.json.
Sheet "reviews", columns:
    shelf | title | author | rating | date_read | isbn | goodreads_link | cover_url | review

- shelf: "currently-reading" or "read" (anything else is ignored)
- cover_url: leave blank to auto-derive from isbn via Open Library covers
- date_read: YYYY-MM-DD (or a real Excel date; both work)

Usage:
    pip install openpyxl   (once)
    python3 scripts/xlsx_to_books.py            # reads books.xlsx -> books.json
    python3 scripts/xlsx_to_books.py mylist.xlsx -o books.json
"""

import argparse
import json
import sys
from datetime import datetime, timezone, date

try:
    from openpyxl import load_workbook
except ImportError:
    print("openpyxl is required: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def norm_date(v) -> str:
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    return str(v).strip() if v else ""


def row_to_book(r: dict) -> dict:
    isbn = str(r.get("isbn") or "").strip()
    cover = str(r.get("cover_url") or "").strip()
    if not cover and isbn:
        cover = f"https://covers.openlibrary.org/b/isbn/{isbn}-M.jpg"
    return {
        "title": str(r.get("title") or "").strip(),
        "author": str(r.get("author") or "").strip(),
        "link": str(r.get("goodreads_link") or "").strip() or "https://www.goodreads.com/",
        "cover": cover,
        "rating": int(r.get("rating") or 0),
        "read_at": norm_date(r.get("date_read")),
        "review": str(r.get("review") or "").strip(),
    }


def main() -> int:
    ap = argparse.ArgumentParser(description="books.xlsx -> books.json")
    ap.add_argument("xlsx_path", nargs="?", default="books.xlsx")
    ap.add_argument("-o", "--out", default="books.json")
    args = ap.parse_args()

    try:
        wb = load_workbook(args.xlsx_path, data_only=True)
    except FileNotFoundError:
        print(f"no such file: {args.xlsx_path}", file=sys.stderr)
        return 1

    ws = wb["reviews"] if "reviews" in wb.sheetnames else wb.active
    header = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

    shelves = {"currently_reading": [], "read": []}
    for row in ws.iter_rows(min_row=2, values_only=True):
        r = dict(zip(header, row))
        if not str(r.get("title") or "").strip():
            continue
        shelf = str(r.get("shelf") or "").strip().lower()
        if shelf == "currently-reading":
            shelves["currently_reading"].append(row_to_book(r))
        elif shelf == "read":
            shelves["read"].append(row_to_book(r))

    shelves["read"].sort(key=lambda b: b["read_at"] or "0000", reverse=True)

    data = {"updated": datetime.now(timezone.utc).isoformat(), **shelves}
    with open(args.out, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)

    print(f"currently reading: {len(shelves['currently_reading'])}, "
          f"read: {len(shelves['read'])} -> wrote {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())